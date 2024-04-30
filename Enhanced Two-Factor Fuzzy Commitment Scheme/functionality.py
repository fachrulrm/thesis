import math
import openpyxl

def xorFunction(input1, input2):
    output = ""

    if len(input1)==len(input2):
        for i in range(len(input1)):
            if input1[i] == input2[i]:
                output = output + "0"
            else:
                output = output + "1"
    else:
        print("Input length must be the same !")

    return output

def hammingDistance(x, y):

    counter = 0
    if len(x) == len(y):
        for i in range(len(x)):
            if x[i] != y[i]:
                counter += 1
    return counter

def normalization(inputArr):
    n = len(inputArr)

    for i in range(n):
        inputArr[i] = inputArr[i]/128

    inputArr = [str(x) for x in inputArr]

    return inputArr

def inputPUFDataOnly(inputArr):
    inputIndex = []

    for i in range(len(inputArr)):
        if inputArr[i] != 0:
            inputIndex.append(i)

    return inputIndex

def mean(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    total = 0

    for i in inputArr:
        total = total + i

    return total/len(inputArr)

def maximum(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    index = 0

    for i in range(len(inputArr)):
        if inputArr[i] > inputArr[index]:
            index = i

    return inputArr[index]

def minimum(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    index = 0

    for i in range(len(inputArr)):
        if inputArr[i] < inputArr[index]:
            index = i

    return inputArr[index]

def standardDeviation(inputArr, mean):
    inputArr = inputPUFDataOnly(inputArr)
    n = len(inputArr)
    sumValue = 0

    for i in range(n):
        sumValue = sumValue + pow((inputArr[i] - mean),2)

    return math.sqrt(sumValue/n)


def decidability(inputHammingDistribution, inputRogueHammingDistribution):
    meanInputHammingDistribution = mean(inputHammingDistribution)
    meanInputRogueHammingDistribution = mean(inputRogueHammingDistribution)

    stdDevInputHammingDistribution = standardDeviation(inputHammingDistribution, meanInputHammingDistribution)
    stdDevInputRogueHammingDistribution = standardDeviation(inputRogueHammingDistribution,
                                                            meanInputRogueHammingDistribution)

    decidabilityValue = (abs(meanInputHammingDistribution - meanInputRogueHammingDistribution)) / (
        math.sqrt((0.5) * (pow(stdDevInputHammingDistribution, 2) + pow(stdDevInputRogueHammingDistribution, 2))))

    if decidabilityValue:
        return decidabilityValue
    else:
        return None


def confusionMatrix(inputHammingDistribution, inputRogueHammingDistribution):
    FAR = 0
    FRR = 0
    TAR = 0
    TRR = 0

    totalHammingDistribution = 0
    totalRogueHammingDistribution = 0

    for x in inputHammingDistribution:
        totalHammingDistribution = totalHammingDistribution + x

    for x in inputRogueHammingDistribution:
        totalRogueHammingDistribution = totalRogueHammingDistribution + x

    maxInputHammingDistribution = maximum(inputHammingDistribution)
    minInputRogueHammingDistribution = minimum(inputRogueHammingDistribution)

    if maxInputHammingDistribution - minInputRogueHammingDistribution > 0:

        if len(inputHammingDistribution) == len(inputRogueHammingDistribution):

            intersect = 0

            for x in range(len(inputHammingDistribution)):
                if (inputHammingDistribution[x] <= inputRogueHammingDistribution[x]) and (
                        x >= minInputRogueHammingDistribution) and (x <= maxInputHammingDistribution):
                    if inputHammingDistribution[x] == inputRogueHammingDistribution[x]:
                        intersect = x
                        break
                    else:
                        intersect = ((x - 1) + (x)) / 2
                        break

            print("Intersect : ", intersect)

            for x in range(len(inputHammingDistribution)):
                if intersect and (x >= intersect) and (x >= minInputRogueHammingDistribution) and (
                        x <= maxInputHammingDistribution):
                    FRR = FRR + inputHammingDistribution[x]

            for x in range(len(inputRogueHammingDistribution)):
                if intersect and x <= intersect and (x >= minInputRogueHammingDistribution) and (
                        x <= maxInputHammingDistribution):
                    FAR = FAR + inputRogueHammingDistribution[x]

            FRR = (FRR / totalHammingDistribution) * 100
            FAR = (FAR / totalRogueHammingDistribution) * 100
            TRR = 100 - FAR
            TAR = 100 - FRR

            return [FRR, FAR, TRR, TAR]
        else:
            return [None, None, None, None]

    else:
        return [None, None, None, None]

def randomness(inputResponseArr):
    randomness = 0
    bitOne     = 0

    for x in range(len(inputResponseArr)): # 0 up to 999
        for y in range(len(inputResponseArr[x])): # 0 up to 127
            if inputResponseArr[x][y] == "1":
                bitOne += 1

    if bitOne:
        randomness = bitOne * 100 / ( len(inputResponseArr) * len(inputResponseArr[0]) )

    # return -1*math.log(max(randomness/100,(1-randomness/100)),2)
    return randomness

def calculateHammingDistance_intraBiometric(inputBiometric : list) -> list:
    hammingDistribution = []

    for i in range(129):  # 0 (All same) up to 128 (All different)
        hammingDistribution.append(0)

    for i in range(len(inputBiometric)):  # 0 sd 999
        for j in range(i + 1, len(inputBiometric)):
            result                       = hammingDistance(inputBiometric[i], inputBiometric[j])
            hammingDistribution[result] += 1

    if hammingDistribution:
        return hammingDistribution
    else:
        return None

def calculateHammingDistance_interBiometric(inputBiometric : list, inputAttackerBiometric : list) -> list:
    hammingDistribution = []

    for i in range(129):  # 0 (All same) up to 128 (All different)
        hammingDistribution.append(0)

    for i in range(len(inputBiometric)):
        for j in range(len(inputAttackerBiometric)):
            result                       = hammingDistance(inputBiometric[i], inputAttackerBiometric[j])
            if result == 2:
                print("Result = 2 happened with "+inputBiometric[i]+" and "+inputAttackerBiometric[j]+", index is "+str(i)+" and "+str(j))
            hammingDistribution[result] += 1

    if hammingDistribution:
        return hammingDistribution
    else:
        return None

def calculateHammingDistance_interBiometric_Genuine(inputBiometric : list, inputAttackerBiometric : list) -> list:
    hammingDistribution = []

    for i in range(129):  # 0 (All same) up to 128 (All different)
        hammingDistribution.append(0)

    for i in range(len(inputBiometric)):
        for j in range(i+1, len(inputAttackerBiometric)):
            result                       = hammingDistance(inputBiometric[i], inputAttackerBiometric[j])
            if result == 2:
                print("Result = 2 happened with "+inputBiometric[i]+" and "+inputAttackerBiometric[j]+", index is "+str(i)+" and "+str(j))
            hammingDistribution[result] += 1

    if hammingDistribution:
        return hammingDistribution
    else:
        return None

def printParameters(inputArr, inputRogueArr, name):
    intraBiometricHammingDistribution = calculateHammingDistance_intraBiometric(inputArr)
    interBiometricHammingDistribution = calculateHammingDistance_interBiometric(inputArr, inputRogueArr)

    print()
    print("======== " + name + " Parameters ========")
    print("INTRA BIOMETRIC PERFORMANCE")
    print()
    print("Mean              : " + str(mean(intraBiometricHammingDistribution)))
    print("Max               : " + str(maximum(intraBiometricHammingDistribution)))
    print("Min               : " + str(minimum(intraBiometricHammingDistribution)))
    print("Stdev             : " + str(standardDeviation(intraBiometricHammingDistribution, mean(intraBiometricHammingDistribution))))
    print()
    print("INTER BIOMETRIC PERFORMANCE")
    print()
    print("Mean              : " + str(mean(interBiometricHammingDistribution)))
    print("Max               : " + str(maximum(interBiometricHammingDistribution)))
    print("Min               : " + str(minimum(interBiometricHammingDistribution)))
    print("Stdev             : " + str(standardDeviation(interBiometricHammingDistribution, mean(interBiometricHammingDistribution))))
    print()
    print("SYSTEM PERFORMANCE")
    print("Decidability      : ", decidability(intraBiometricHammingDistribution, interBiometricHammingDistribution))
    print("Randomness        : ", randomness(inputArr))
    outputConfusionMatrix = confusionMatrix(intraBiometricHammingDistribution, interBiometricHammingDistribution)
    if outputConfusionMatrix:
        if len(outputConfusionMatrix) == 4:
            print("FRR               : ", outputConfusionMatrix[0])
            print("FAR               : ", outputConfusionMatrix[1])
            print("TRR               : ", outputConfusionMatrix[2])
            print("TAR               : ", outputConfusionMatrix[3])
    print("======== END ========")
    print()

def printParameters_Genuine(inputArr, inputRogueArr, name):
    intraBiometricHammingDistribution = calculateHammingDistance_intraBiometric(inputArr)
    interBiometricHammingDistribution = calculateHammingDistance_interBiometric_Genuine(inputArr, inputRogueArr)

    print()
    print("======== " + name + " Parameters ========")
    print("INTRA BIOMETRIC PERFORMANCE")
    print()
    print("Mean              : " + str(mean(intraBiometricHammingDistribution)))
    print("Max               : " + str(maximum(intraBiometricHammingDistribution)))
    print("Min               : " + str(minimum(intraBiometricHammingDistribution)))
    print("Stdev             : " + str(standardDeviation(intraBiometricHammingDistribution, mean(intraBiometricHammingDistribution))))
    print()
    print("INTER BIOMETRIC PERFORMANCE")
    print()
    print("Mean              : " + str(mean(interBiometricHammingDistribution)))
    print("Max               : " + str(maximum(interBiometricHammingDistribution)))
    print("Min               : " + str(minimum(interBiometricHammingDistribution)))
    print("Stdev             : " + str(standardDeviation(interBiometricHammingDistribution, mean(interBiometricHammingDistribution))))
    print()
    print("SYSTEM PERFORMANCE")
    print("Decidability      : ", decidability(intraBiometricHammingDistribution, interBiometricHammingDistribution))
    print("Randomness        : ", randomness(inputArr))
    outputConfusionMatrix = confusionMatrix(intraBiometricHammingDistribution, interBiometricHammingDistribution)
    if outputConfusionMatrix:
        if len(outputConfusionMatrix) == 4:
            print("FRR               : ", outputConfusionMatrix[0])
            print("FAR               : ", outputConfusionMatrix[1])
            print("TRR               : ", outputConfusionMatrix[2])
            print("TAR               : ", outputConfusionMatrix[3])
    print("======== END ========")
    print()

def importDataInternalNoisySource(filePath, CRPCount):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active

    responsesArr = []

    for i in range(CRPCount):
        responsesString      = sheet.cell(row=i + 2, column=2).value
        responsesArr.append(responsesString)

    if responsesArr:
        return responsesArr
    else:
        return None

def importDataRogueInternalNoisySource(filePath, CRPCount):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active

    responsesArr = []

    for i in range(CRPCount):
        responsesString      = sheet.cell(row=i + 2, column=3).value
        responsesArr.append(responsesString)

    if responsesArr:
        return responsesArr
    else:
        return None

def convertingExternalNoisySource_String_to_Binary(inputData):
    temperature     = inputData[0]
    longitude       = inputData[1]
    lattitude       = inputData[2]
    altitude        = inputData[3]

    # Parsing longitude & lattitude
    longitude = longitude[longitude.find(".") + 2:longitude.find(".") + 7]
    lattitude = lattitude[lattitude.find(".") + 2:lattitude.find(".") + 7]

    externalNoisySource = temperature + longitude + lattitude + altitude

    externalNoisySourceBIN = str(bin(int.from_bytes(externalNoisySource.encode(), "big")))
    externalNoisySourceBIN = externalNoisySourceBIN[2:len(externalNoisySourceBIN)]

    if len(externalNoisySourceBIN) < 128:
        counterZero = 128 - len(externalNoisySourceBIN)
        for i in range(counterZero):
            externalNoisySourceBIN = "0" + externalNoisySourceBIN

    if len(externalNoisySourceBIN)==128:
        return externalNoisySourceBIN
    else:
        return None

def integratingBiometrics(internalNoisySource : list, externalNoisySource : str) -> list:
    newBiometric    = []

    for i in internalNoisySource:
        if len(i) == len(externalNoisySource):
            newBiometric.append(xorFunction(i, externalNoisySource))

    if len(newBiometric) == len(internalNoisySource):
        return newBiometric
    else:
        return None

def xLabelForGraph(minValue : int, maxValue : int) -> list:
    xLabel = []

    for i in range(minValue, maxValue+1):
        xLabel.append(i)

    if xLabel:
        return xLabel
    else:
        return None

def generatingHelperBitsFromVarianceDataset(inputBiometric : list, inputCodeword : str) -> list:
    outputArray = []

    if inputBiometric and len(inputBiometric) == 100:
        for dataset in inputBiometric:
            if dataset and len(dataset) == 1000:
                tempArray = []

                for i in dataset:
                    if len(i) == len(inputCodeword):
                        tempArray.append(xorFunction(i, inputCodeword))
                    else:
                        break
                        return None

                if tempArray:
                    outputArray.append(tempArray)
                else:
                    break
                    return None
            else:
                break
                return None
    else:
        return None

    if outputArray and len(outputArray) == len(inputBiometric):
        return outputArray
    else:
        return None

def generatingReceivedCodewordFromVarianceDataset(inputHelperBitsFromVarianceDataset : list, inputAuthenticBiometric : list) -> list:
    outputArray = []

    if inputHelperBitsFromVarianceDataset and len(inputHelperBitsFromVarianceDataset) == 100:
        for datasetHelperBits in inputHelperBitsFromVarianceDataset:
            if datasetHelperBits and len(datasetHelperBits) == len(inputAuthenticBiometric):
                tempArray = []

                for i in range(len(datasetHelperBits)):
                    if len(datasetHelperBits[i]) == len(inputAuthenticBiometric[i]):
                        tempArray.append(xorFunction(datasetHelperBits[i], inputAuthenticBiometric[i]))
                    else:
                        break
                        return None

                if tempArray:
                    outputArray.append(tempArray)
                else:
                    break
                    return None

            else:
                break
                return None

    else:
        return None

    if outputArray and len(outputArray) == len(inputHelperBitsFromVarianceDataset):
        return outputArray
    else:
        return None

def savingToExcel(filePath : str, inputExternalNoisySource_plainText : list, inputKRR : list):
    if filePath and inputExternalNoisySource_plainText and inputKRR:
        wb              = openpyxl.Workbook()
        writeExcelSheet = wb.active

        cell            = writeExcelSheet.cell(row=1, column=1)
        cell.value      = "External Noisy Source (Plain Text)"

        for i in range(len(inputExternalNoisySource_plainText)):  # 0 up to 999
            cell        = writeExcelSheet.cell(row=i + 2, column=1)
            cell.value  = ("Temp : "+inputExternalNoisySource_plainText[i][0] + ", Longitude : "+inputExternalNoisySource_plainText[i][1] +
                           ", Lattitude : "+inputExternalNoisySource_plainText[i][2] + ", Altitude :"+inputExternalNoisySource_plainText[i][3])

        cell            = writeExcelSheet.cell(row=1, column=2)
        cell.value      = "External Noisy Source (Converted to Binary)"

        for i in range(len(inputExternalNoisySource_plainText)):  # 0 up to 999
            cell        = writeExcelSheet.cell(row=i + 2, column=2)
            cell.value  = convertingExternalNoisySource_String_to_Binary(inputExternalNoisySource_plainText[i])

        cell            = writeExcelSheet.cell(row=1, column=3)
        cell.value      = "Number of Bits for Key Recovery"

        for i in range(len(inputExternalNoisySource_plainText)):  # 0 up to 999
            cell        = writeExcelSheet.cell(row=i + 2, column=3)
            cell.value  = inputKRR[i]

        wb.save(filePath)
    else:
        return None

def readExcelColumn(filePath : str, column : int, totalData : int) -> list:
    outputArray     = []
    book            = openpyxl.load_workbook(filePath)
    sheet           = book.active

    for i in range(totalData):
        outputArray.append(sheet.cell(row=i + 2, column=column).value)

    if outputArray:
        return outputArray
    else:
        return None

def calculatingKRRfromKRRsDataset(inputKRRDataset : list, inputXLabel : list) -> list:
    outputArray = []

    if inputKRRDataset and inputXLabel:
        for i in inputXLabel:
            counter = 0

            for j in inputKRRDataset:
                if j <= i:
                    counter += 1

            outputArray.append(counter)

        if outputArray:
            return outputArray
        else:
            return None
    else:
        return None

def checkAttackerKRR(inputKRR : list) -> list:
    if inputKRR:
        index = 0

        for i in range(len(inputKRR)):
            if inputKRR[i] > 0:
                index = i
                break

        return index - 1
    else:
        return None