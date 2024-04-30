# CRPs sets
CRPCount = 1000 # <----------------------------- Change this according to your CRP count needs

# Initialization - Read Excel Document
import openpyxl
book = openpyxl.load_workbook('Datasets.xlsx')
sheet = book.active

# To create excel documents (write pattern that shows the different between PUF responses)
wb              = openpyxl.Workbook()
writeExcelSheet = wb.active

# Getting PUF Responses Data from Excel
challengesArr      = []
responsesArr      = []
responsesRogueArr = []

for i in range(CRPCount):
    challengesString     = sheet.cell(row=i + 2, column=1).value
    responsesString      = sheet.cell(row=i + 2, column=2).value
    responsesRogueString = sheet.cell(row=i + 2, column=3).value

    challengesArr.append(challengesString)
    responsesArr.append(responsesString)
    responsesRogueArr.append(responsesRogueString)

# ==============================================================================================================================================
# Calculating Hamming Distance
class Solution(object):
    def hammingDistance(self, x, y, inputContributingArr):

        counter = 0
        if len(x) == len(y):
            for i in range(len(x)):
                if x[i] != y[i]:
                    inputContributingArr[i] += 1
                    counter += 1
        return counter

objHammingDistance = Solution()

hammingDistribution = [] # i.e. hammingDistribution[0] = 10 means that there are 10 cases that have hamming distance of 0
hammingLabel        = [] # i.e. for plot labelling purposes
contributingBit     = [] # i.e first bit is the most contributing bit for HD calculation (most often to be different)
contributingBitLabel= [] # i.e. for plot labelling purposes

for i in range(129): # 129 because hamming distance calculation will result from 0 (All same) up to 128 (All different). Thus, total space needed in array is 129
    hammingDistribution.append(0)
    hammingLabel.append(i)

for i in range(128): # bit 0 up to 127
    contributingBit.append(0)
    contributingBitLabel.append(str(i))

for i in range(len(responsesArr)):  # 0 sd 999
    for j in range(i + 1, len(responsesArr)):
        result = Solution().hammingDistance(responsesArr[i], responsesArr[j], contributingBit)
        hammingDistribution[result] += 1

# ==============================================================================================================================================
# Normalization

def normalization(inputArr):
    n = len(inputArr)

    for i in range(n):
        inputArr[i] = inputArr[i]/128

    inputArr = [str(x) for x in inputArr]

    return inputArr

hammingLabel = normalization(hammingLabel)

# ==============================================================================================================================================
# Parameter Calculation
import math

def inputPUFDataOnly(inputArr):
    inputIndex = []

    for i in range(len(inputArr)):
        if inputArr[i] != 0:
            inputIndex.append(i)

    return inputIndex

def mean(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    total = 0

    for i in inputArr:
        total = total + i

    return total/len(inputArr)

def maximum(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    index = 0

    for i in range(len(inputArr)):
        if inputArr[i] > inputArr[index]:
            index = i

    return inputArr[index]

def minimum(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    index = 0

    for i in range(len(inputArr)):
        if inputArr[i] < inputArr[index]:
            index = i

    return inputArr[index]

def standardDeviation(inputArr, mean):
    inputArr = inputPUFDataOnly(inputArr)
    n = len(inputArr)
    sumValue = 0

    for i in range(n):
        sumValue = sumValue + pow((inputArr[i] - mean),2)

    return math.sqrt(sumValue/n)

def printParameters(inputArr):
    print()
    print("======== PUF Parameters ========")
    print("Mean : " + str(mean(inputArr)))
    print("Max  : " + str(maximum(inputArr)))
    print("Min  : " + str(minimum(inputArr)))
    print("Stdev: " + str(standardDeviation(inputArr, mean(inputArr))))
    print()

printParameters(hammingDistribution)

# ==============================================================================================================================================
# Creating the line plot for contributing bit

import pandas as pd
import matplotlib.pyplot as plt

data_dict = {'Bit Position': contributingBitLabel,
			'Differences': contributingBit
			}

df = pd.DataFrame(data_dict)

# Checking first 5 data by default in dictionary
# print(df.head())

ax = plt.gca()
df.plot(kind='line',
		x='Bit Position',
		y='Differences',
		color='red',
        ax=ax)

plt.title('LinePlots')
plt.show()

# ==============================================================================================================================================
# Dataset Processing - Truncated Bits 4, 8, 12, 16 bits

uniformedBits = [4, 8, 12, 16]

def datasetProcessing(inputUniformedBit, inputContributingBit, inputResponseArr):
    processedBits            = []
    tempInputContributingBit = []

    for x in inputContributingBit:
        tempInputContributingBit.append(x)

    for x in range(inputUniformedBit): # 0 up to 3 / 7 / 11 / 15
        index = tempInputContributingBit.index(max(tempInputContributingBit)) # index that contain maximum value
        processedBits.append(index)
        tempInputContributingBit[index] = 0

    print(processedBits)

    modifiedResponseArr      = []

    for x in range(len(inputResponseArr)): # 0 up to 999
        modifiedResponse      = ""

        for y in range(len(inputResponseArr[x])): # 0 up to 127

            if y in processedBits:
                modifiedResponse = modifiedResponse + "0" # Chenge this particular to '0' value
            else:
                modifiedResponse      = modifiedResponse + inputResponseArr[x][y]

        if modifiedResponse:
            modifiedResponseArr.append(modifiedResponse)

        modifiedResponse      = ""

    return [modifiedResponseArr]

def contributingBitCalculation(inputResponseArr):
    inputContributingBits = list()

    for x in range(len(inputResponseArr[0])):
        inputContributingBits.append(0)

    for x in range(len(inputResponseArr)): # 0 up to 999
        for y in range(x + 1, len(inputResponseArr)):
            if len(inputResponseArr[x]) == len(inputResponseArr[y]):
                for z in range(len(inputResponseArr[x])): # 0 up to 127 / 123 / 119 / 115
                    if inputResponseArr[x][z] != inputResponseArr[y][z]:
                        inputContributingBits[z] += 1

    return inputContributingBits

# Creating New Uniformed Response Arr
responseArr_4  = list()
responseArr_8  = list()
responseArr_12 = list()
responseArr_16 = list()

processedResponseArr = []

for i in uniformedBits:
    if i == 4:
        output_datasetProcessing = datasetProcessing(i, contributingBit, responsesArr)

        responseArr_4 = output_datasetProcessing[0]
        processedResponseArr.append(responseArr_4)
    elif i == 8:
        output_datasetProcessing = datasetProcessing(i, contributingBit, responsesArr)

        responseArr_8 = output_datasetProcessing[0]
        processedResponseArr.append(responseArr_8)
    elif i == 12:
        output_datasetProcessing = datasetProcessing(i, contributingBit, responsesArr)

        responseArr_12 = output_datasetProcessing[0]
        processedResponseArr.append(responseArr_12)
    elif i == 16:
        output_datasetProcessing = datasetProcessing(i, contributingBit, responsesArr)

        responseArr_16 = output_datasetProcessing[0]
        processedResponseArr.append(responseArr_16)

# Calculating New Contributing Bits
contributingBit_4  = list()
contributingBit_8  = list()
contributingBit_12 = list()
contributingBit_16 = list()

processedcontributingBit = []

if len(processedResponseArr) == 4:
    for i in range(len(processedResponseArr)):
        if i == 0:
            contributingBit_4 = contributingBitCalculation(processedResponseArr[i])
            processedcontributingBit.append(contributingBit_4)
        elif i == 1:
            contributingBit_8 = contributingBitCalculation(processedResponseArr[i])
            processedcontributingBit.append(contributingBit_8)
        elif i == 2:
            contributingBit_12 = contributingBitCalculation(processedResponseArr[i])
            processedcontributingBit.append(contributingBit_12)
        elif i == 3:
            contributingBit_16 = contributingBitCalculation(processedResponseArr[i])
            processedcontributingBit.append(contributingBit_16)

# ==============================================================================================================================================
# Creating the line plot for contributing bit

def plotContributingBit(inputContributingBitArr):
    if len(inputContributingBitArr) != len(contributingBitLabel):
        while len(inputContributingBitArr) < len(contributingBitLabel):
            inputContributingBitArr.append(None)

    data_dict = {'Bit Position': contributingBitLabel,
                 'Occurrence of different bits': inputContributingBitArr
                 }

    df = pd.DataFrame(data_dict)

    ax = plt.gca()
    df.plot(kind='line',
            x='Bit Position',
            y='Occurrence of different bits',
            color='red',
            ax=ax)

    plt.title('LinePlots')
    plt.show()

if len(processedcontributingBit) == 4:
    for i in processedcontributingBit:
        plotContributingBit(i)

# ==============================================================================================================================================
# Writing the output to Excel Document

def saveToExcel(inputUniformedBits, inputChallengeArr, inputResponseArr, inputResponseRogueArr):
    cell = writeExcelSheet.cell(row=1, column=1)
    cell.value = "Challenge Bits"

    cell = writeExcelSheet.cell(row=1, column=2)
    cell.value = "Genuine PUF Response"

    cell = writeExcelSheet.cell(row=1, column=3)
    cell.value = "Rogue PUF Response"

    if (len(inputChallengeArr) == len(inputResponseArr)) and (len(inputChallengeArr) == len(inputResponseRogueArr)):
        # Writing Challenge Bits
        for i in range(len(inputChallengeArr)): # 0 up to 999
            cell = writeExcelSheet.cell(row=i + 2, column=1)
            cell.value = inputChallengeArr[i]

        # Writing Responses Bits
        for i in range(len(inputResponseArr)): # 0 up to 999
            cell = writeExcelSheet.cell(row=i + 2, column=2)
            cell.value = inputResponseArr[i]

        # Writing Responses Rogue Bits
        for i in range(len(inputResponseRogueArr)): # 0 up to 999
            cell = writeExcelSheet.cell(row=i + 2, column=3)
            cell.value = inputResponseRogueArr[i]

        wb.save("/Users/fachrulreizamedina/Python/Thesis/1. Implementation - Old Scheme, New Code/Datasets_Uniformed_{}.xlsx".format(inputUniformedBits))

for i in range(len(uniformedBits)): # 0 sd 3
    saveToExcel(uniformedBits[i], challengesArr, processedResponseArr[i], responsesRogueArr)