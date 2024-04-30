# CRPs sets
CRPCount = 1000 # <----------------------------- Change this according to your CRP count needs

import openpyxl
import math
import pandas as pd
import matplotlib.pyplot as plt

def readExcel(inputTruncatedBits, inputDatasetLength):
    # Initialization - Read Excel Document
    book  = openpyxl.load_workbook('Datasets_Truncated_{}.xlsx'.format(inputTruncatedBits))
    sheet = book.active

    outputChallengesArr     = list()
    outputResponsesArr      = list()
    outputResponsesRogueArr = list()

    for x in range(inputDatasetLength):
        outputChallengesString = sheet.cell(row=x + 2, column=1).value
        outputResponsesString = sheet.cell(row=x + 2, column=2).value
        outputResponsesRogueString = sheet.cell(row=x + 2, column=3).value

        outputChallengesArr.append(outputChallengesString)
        outputResponsesArr.append(outputResponsesString)
        outputResponsesRogueArr.append(outputResponsesRogueString)

    if (len(outputChallengesArr) == len(outputResponsesArr)) and (len(outputChallengesArr) == len(outputResponsesRogueArr)) and outputChallengesArr and outputResponsesArr and outputResponsesRogueArr:
        return [outputChallengesArr, outputResponsesArr, outputResponsesRogueArr]
    else:
        return [None, None, None]

def normalization(inputArr):
    n = len(inputArr)

    for i in range(n):
        inputArr[i] = inputArr[i]/128

    inputArr = [str(x) for x in inputArr]

    return inputArr

def hammingDistance(x, y):
    counter = 0

    if len(x) == len(y):
        for i in range(len(x)):
            if x[i] != y[i]:
                counter += 1

    return counter

def calculateHammingDistribution_External(inputResponseArr, inputResponseRogueArr):
    if inputResponseArr and inputResponseRogueArr:
        if len(inputResponseArr[0]) == len(inputResponseRogueArr[0]):

            hammingDistribution = []  # i.e. hammingDistribution[0] = 10 means that there are 10 cases that have hamming distance of 0
            hammingLabel        = []  # i.e. for plot labelling purposes

            for x in range(len(inputResponseArr[0]) + 1):  # 0 up to 128 / 124 / 120 / 116 / 112, +1 because it can be all same (0) or all different (128 / 124 / 120 / 116 / 112)
                hammingDistribution.append(0)
                hammingLabel.append(x)

            for x in range(len(inputResponseArr)): # 0 sd 999
                for y in range(len(inputResponseRogueArr)): # 0 sd 999
                    result                       = hammingDistance(inputResponseArr[x], inputResponseRogueArr[y])
                    hammingDistribution[result] += 1

            if hammingDistribution and hammingLabel:
                hammingLabel = normalization(hammingLabel)

                return [hammingDistribution, hammingLabel]
            else:
                return [None, None]

        else:
            return [None, None]

    else:
        return [None, None]

def calculateHammingDistribution_Internal(inputResponseArr):
    if inputResponseArr:

        hammingDistribution = []  # i.e. hammingDistribution[0] = 10 means that there are 10 cases that have hamming distance of 0
        hammingLabel        = []  # i.e. for plot labelling purposes

        for x in range(len(inputResponseArr[0]) + 1):  # 0 up to 128 / 124 / 120 / 116 / 112, +1 because it can be all same (0) or all different (128 / 124 / 120 / 116 / 112)
            hammingDistribution.append(0)
            hammingLabel.append(x)

        for x in range(len(inputResponseArr)): # 0 sd 999
            for y in range(x + 1, len(inputResponseArr)): # up to 999
                result                       = hammingDistance(inputResponseArr[x], inputResponseArr[y])
                hammingDistribution[result] += 1

        if hammingDistribution and hammingLabel:
            hammingLabel = normalization(hammingLabel)

            return [hammingDistribution, hammingLabel]
        else:
            return [None, None]

    else:
        return [None, None]

def inputPUFDataOnly(inputArr):
    inputIndex = []

    for i in range(len(inputArr)):
        if inputArr[i] != 0:
            inputIndex.append(i)

    return inputIndex

def mean(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    total = 0

    for i in inputArr:
        total = total + i

    return total/len(inputArr)

def maximum(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    index = 0

    for i in range(len(inputArr)):
        if inputArr[i] > inputArr[index]:
            index = i

    return inputArr[index]

def minimum(inputArr):
    inputArr = inputPUFDataOnly(inputArr)
    index = 0

    for i in range(len(inputArr)):
        if inputArr[i] < inputArr[index]:
            index = i

    return inputArr[index]

def standardDeviation(inputArr, mean):
    inputArr = inputPUFDataOnly(inputArr)
    n = len(inputArr)
    sumValue = 0

    for i in range(n):
        sumValue = sumValue + pow((inputArr[i] - mean),2)

    return math.sqrt(sumValue/n)

def decidability(inputHammingDistribution, inputRogueHammingDistribution):
    meanInputHammingDistribution      = mean(inputHammingDistribution)
    meanInputRogueHammingDistribution = mean(inputRogueHammingDistribution)

    stdDevInputHammingDistribution      = standardDeviation(inputHammingDistribution, meanInputHammingDistribution)
    stdDevInputRogueHammingDistribution = standardDeviation(inputRogueHammingDistribution, meanInputRogueHammingDistribution)

    decidabilityValue = ( abs(meanInputHammingDistribution - meanInputRogueHammingDistribution) ) / ( math.sqrt( (0.5)*( pow(stdDevInputHammingDistribution, 2) + pow(stdDevInputRogueHammingDistribution, 2) ) ) )

    if decidabilityValue:
        return decidabilityValue*100
    else:
        return None

def confusionMatrix(inputHammingDistribution, inputRogueHammingDistribution):
    FAR     = 0
    FRR     = 0
    TAR     = 0
    TRR     = 0

    totalHammingDistribution        = 0
    totalRogueHammingDistribution   = 0

    for x in inputHammingDistribution:
        totalHammingDistribution = totalHammingDistribution + x

    for x in inputRogueHammingDistribution:
        totalRogueHammingDistribution = totalRogueHammingDistribution + x

    maxInputHammingDistribution      = maximum(inputHammingDistribution)
    minInputRogueHammingDistribution = minimum(inputRogueHammingDistribution)

    if maxInputHammingDistribution - minInputRogueHammingDistribution > 0:

        if len(inputHammingDistribution) == len(inputRogueHammingDistribution):

            intersect = 0

            for x in range(len(inputHammingDistribution)):
                if (inputHammingDistribution[x] <= inputRogueHammingDistribution[x]) and (x >= minInputRogueHammingDistribution) and (x <= maxInputHammingDistribution):
                    if inputHammingDistribution[x] == inputRogueHammingDistribution[x]:
                        intersect = x
                        break
                    else:
                        intersect = ((x - 1) + (x)) / 2
                        break

            print("Intersect : ", intersect)

            for x in range(len(inputHammingDistribution)):
                if intersect and (x >= intersect) and (x >= minInputRogueHammingDistribution) and (x <= maxInputHammingDistribution):
                    FRR = FRR + inputHammingDistribution[x]

            for x in range(len(inputRogueHammingDistribution)):
                if intersect and x <= intersect and (x >= minInputRogueHammingDistribution) and (x <= maxInputHammingDistribution):
                    FAR = FAR + inputRogueHammingDistribution[x]

            FRR = (FRR/totalHammingDistribution)*100
            FAR = (FAR/totalRogueHammingDistribution)*100
            TRR = 100 - FAR
            TAR = 100 - FRR

            return [FRR, FAR, TRR, TAR]
        else:
            return [None, None, None, None]

    else:
        return [None, None, None, None]

def randomness(inputResponseArr):
    randomness = 0
    bitOne     = 0

    for x in range(len(inputResponseArr)): # 0 up to 999
        for y in range(len(inputResponseArr[x])): # 0 up to 127
            if inputResponseArr[x][y] == "1":
                bitOne += 1

    if bitOne:
        randomness = bitOne * 100 / ( len(inputResponseArr) * len(inputResponseArr[0]) )

    return randomness

def printParameters(inputArr):
    print()
    print("======== Parameters ========")
    print("Mean : " + str(mean(inputArr)))
    print("Max  : " + str(maximum(inputArr)))
    print("Min  : " + str(minimum(inputArr)))
    print("Stdev: " + str(standardDeviation(inputArr, mean(inputArr))))
    print()\

def plotHD(inputHDLabel, inputHD, inputRogueHD):
    data_dict = {'HD': inputHDLabel,
                 'Intra PUF HD': inputHD,
                 'Inter PUF HD': inputRogueHD
                 }

    df = pd.DataFrame(data_dict)

    ax = plt.gca()
    df.plot(kind='line',
            x='HD',
            y='Intra PUF HD',
            color='green',
            ax=ax)

    df.plot(kind='line',
            x='HD',
            y='Inter PUF HD',
            color='red',
            ax=ax)

    plt.title('LinePlots')
    plt.show()

# ==============================================================================================================================================
# Main Function

truncatedBits       = [0, 4, 8, 12, 16]

challengesArr       = []
responsesArr        = []
responsesRogueArr   = []

hammingDistribution         = []
hammingDistributionLabel    = []

rogueHammingDistribution        = []
rogueHammingDistributionLabel   = []

for i in truncatedBits:
    # Getting PUF Responses Data from Excel
    outputReadExcel = readExcel(i, CRPCount)

    if outputReadExcel:
        if len(outputReadExcel) == 3:
            challengesArr       = outputReadExcel[0]
            responsesArr        = outputReadExcel[1]
            responsesRogueArr   = outputReadExcel[2]

            print(len(responsesArr[0]))

    # Calculate HD
    if responsesArr and responsesRogueArr:
        # Calculate External HD
        outputCalculateHammingDistribution_External = calculateHammingDistribution_External(responsesArr, responsesRogueArr)

        rogueHammingDistribution        = outputCalculateHammingDistribution_External[0]
        rogueHammingDistributionLabel   = outputCalculateHammingDistribution_External[1]

        # Calculate Internal HD
        outputCalculateHammingDistribution_Internal = calculateHammingDistribution_Internal(responsesArr)

        hammingDistribution         = outputCalculateHammingDistribution_Internal[0]
        hammingDistributionLabel    = outputCalculateHammingDistribution_Internal[1]

        print("============== START ==================")
        print("Truncated Bits    : ", i)
        print("Decidability      : ", decidability(hammingDistribution, rogueHammingDistribution))
        print("Randomness        : ", randomness(responsesArr))

        outputConfusionMatrix = confusionMatrix(hammingDistribution, rogueHammingDistribution)
        if outputConfusionMatrix:
            if len(outputConfusionMatrix) == 4:
                print("FRR : ", outputConfusionMatrix[0])
                print("FAR : ", outputConfusionMatrix[1])
                print("TRR : ", outputConfusionMatrix[2])
                print("TAR : ", outputConfusionMatrix[3])
        print()
        print("Intra PUF HD Parameters")
        printParameters(hammingDistribution)
        print()
        print("Inter PUF HD Parameters")
        printParameters(rogueHammingDistribution)
        print("==============  END  ==================")

    # Plot HD
    if hammingDistribution and rogueHammingDistribution:
        if (len(hammingDistributionLabel) == len(rogueHammingDistributionLabel)) and (len(hammingDistributionLabel) == len(hammingDistribution)) and (len(hammingDistributionLabel) == len(rogueHammingDistribution)):
            plotHD(hammingDistributionLabel, hammingDistribution, rogueHammingDistribution)